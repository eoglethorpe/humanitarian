"""Module for extracting data from dropbox and aggregating"""
#TODO: better way for specifyign which folder to pull from (latest?)
##how to handle names?
##pull from rules based text file
##log put on dbox, make better
##test individual cleaning
##mvoe current logs to an 'old' file 
##put in run params: test, exclude, folder location?
##change to pull down all WSs at same time as opposed to iterating  
##get rid of spcifycing column for uid, should just be one after header
##logs output to local or dbox
##do quick reading method at first for faster run time
import datetime

import dropbox
import clean
import os
import cStringIO
import re
from openpyxl import load_workbook
from openpyxl.cell import column_index_from_string
from openpyxl import Workbook
import openpyxl.writer.excel as wrtex
import time
import click
from os import listdir
from os.path import isfile, join
from dateutil.parser import parse

#dropbox setup
db_access = os.environ['db_access']
client = dropbox.client.DropboxClient(db_access)

@click.command()
@click.option('--act', help='what action will we perform', type = click.Choice(['cons','clean']))
@click.option('--src', help='local files or on dropbox?', type = click.Choice(['db','local']))
@click.option('--path', help='file path (pull all xls or xlsx files in it')
@click.option('--db', help='db file if consolidating')
@click.option('--test', help='are we testing?', is_flag = True)

def iterate_reports(act, src, path, db, test):
    """cycle through all reports contained in dbox directory"""

    file_list = get_file_list(path, src)

    if test == True:
        file_list = [file_list[0]]

    for v in file_list:
        print v

    #pull down all workbooks
    wbs = []
    for f in file_list:
        #pull down workbook from specified directory
        print "Pulling: " + f
        wb_current = pull_wb(f, src)

        #check to see if properly formatted
        if wb_format(wb_current):
            print 'Appending: ' + f
            wbs.append((wb_current,f))

        else:
            #put in malformatted folder
            print 'Malformatted ' + f
            send_path = path + '/old_format_or_incorrect/' + f.rsplit('/', 1)[1]
            send_wb(send_path, wb_current, src)

    #clean or consolidate
    if act == 'clean':
        #clean workboooks
        for wb in wbs:
            print "Cleaning: " + wb[1]
            clean_file(wb[0], wb[1], src)

    elif act == 'cons':
        #consolidate
        to_send = consolidate(pull_wb(db, src).get_sheet_by_name('Database'), wbs, 'V')
        send_wb(path + 'merged.xlsx', to_send, src)

def clean_exclude(act, file_list):
    if act =='clean':
        new_list = []
        for v in file_list:
            if 'C-' in v or 'C -' in v:
                new_list.append(v)
            else:
                print 'Excluding: ' + v
        file_list = new_list

def consolidate(baseline, wbs, key_col):
    """consolidate baseline data and worksheets into one sheet 
        and remove duplicates"""

    print 'in consolidate'

    cons_wb = Workbook()
    cons = cons_wb.active
    cons.title = 'Consolidated'
    cons.append(get_values(baseline.rows[0]) + ['UID'])

    print 'step1'

    merge_sheets = []
    #pull out desired worksheet
    for wb in wbs:
        print 'step2'
        merge_sheets.append(wb[0].get_sheet_by_name('Distributions'))

    print 'step3'

    #read in baseline db and new WSs into dictionaries
    base_dict = {}
    merge_sheets_dict = {}
    key_loc = column_index_from_string(key_col)-1

    print 'step4'


    #add in UID value for sheets
    #there is a problem where imaginary columns are being returned by .rows, so we find the
    #max value of the header and only look for values of that row that extend out to that column
    for wb in merge_sheets:
        print 'step5'
        max_val = len(wb.rows[0])
        for r in wb.rows[1:]:
            wb[key_col + xstr(r[0].row)] = get_uid(r[0:max_val], wb)


    print 'in 1'

    #add UID for baseline
    max_val = len(baseline.rows[0])
    for r in baseline.rows[1:]:
        baseline[key_col + xstr(r[0].row)] = get_uid(r[0:max_val], baseline)
    print 'in 2'

    #must check if we already have duplicate keys in dictionary or in sheets
    #if we do: take the entry with greater status, comp date then started date

    #add baseline to dict
    dup_count = 0
    for r in baseline.rows[1:]:
        uid = get_uid(r[0:max_val], baseline)
        if base_dict.has_key(uid):
            if not keep_dict(r, base_dict[uid], baseline):
                base_dict[xstr(r[key_loc].value)] = get_values(r)
        else:
                base_dict[xstr(r[key_loc].value)] = get_values(r)

    print 'Found conflicting entries in baseline: ' + str(dup_count)
    print 'in 3'

    #merge sheets into a dict
    dup_count = 0
    ongoing_count = 0
    for wb in merge_sheets:
        for r in wb.rows[1:]:
            uid = get_uid(r[0:max_val], wb)
            if merge_sheets_dict.has_key(uid):
                if not none_row(uid):
                    dup_count+=1
                    print uid
                if not keep_dict(r, merge_sheets_dict[uid], wb):
                    merge_sheets_dict[xstr(r[key_loc].value)] = get_values(r)
            else:
                merge_sheets_dict[xstr(r[key_loc].value)] = get_values(r)

        print 'Found conflicting entries in' + wb['A2'].value + ' : ' + str(dup_count)
        ongoing_count+=dup_count
        dup_count = 0

    print 'Found conflicting entires in all sheets: ' + str(ongoing_count)
    print 'in 4'


    dup_count = 0
    print "**Duplicates**"
    #go through baseline and remove dups
    for k in base_dict.keys():
        if k in merge_sheets_dict.keys():
            print k
            base_dict.pop(k)
            dup_count+=1

    print 'in 5'
    print 'dupcount: ' + str(dup_count)

    #now, add baseline and then wbs to cons
    for v in base_dict.values():
        cons.append(v)
    print 'in 6'

    for v in merge_sheets_dict.values():
        cons.append(v)
    
    return cons_wb

def none_row(val):
    i = (val+val).find(val, 1, -1)
    return None if i == -1 else val[:i]

def keep_dict(row, existing, ws):
    """given conflicting dict entries, return True if current val we should keep
        ...heirarcy of greater status, comp date then started date"""

    status_loc = column_index_from_string(find_in_header(ws, 'Activity Status'))-1
    comp_loc = column_index_from_string(find_in_header(ws, 'Completion Date\n (Actual or Planned)'))-1
    start_loc = column_index_from_string(find_in_header(ws, 'Start date \n(Actual or Planned)'))-1

    #check status
    status = ['Complete', 'Ongoing', 'Plan']
    r_ind = 0
    e_ind = 0

    c=0
    for v in status:
	if row[status_loc].value is None:
	    row[status_loc].value = 'Planned'
	if existing[status_loc] == 'None':
	    existing[status_loc] = 'Planned'

        
	if v in row[status_loc].value:
            r_ind = c
        if v in existing[status_loc]:
            e_ind = c
        c+=1

    #check index
    to_return = False
    if r_ind > e_ind:
        to_return = True

    #check comp date
    if type(row[comp_loc].value) is not datetime.date and type(row[comp_loc].value) is not datetime.datetime and row[comp_loc].value is not None:
        r_v = parse(row[comp_loc].value)
    else:
        r_v = row[comp_loc].value

    if type(existing[comp_loc]) is not datetime.date and type(existing[comp_loc]) is not datetime.datetime and existing[comp_loc] != 'None':
	e_v = parse(existing[comp_loc])
    else:
        e_v = existing[comp_loc]

    if row[comp_loc].value is not None and existing[comp_loc] != 'None':
        if r_v < e_v:
            to_return = True

    #check start date

    if type(row[start_loc].value) is not datetime.date and type(row[start_loc].value) is not datetime.datetime and row[start_loc].value is not None:
        r_v = parse(row[start_loc].value)
    else:
        r_v = row[start_loc].value

    if type(existing[start_loc]) is not datetime.date and type(existing[start_loc]) is not datetime.datetime and existing[start_loc] != 'None':
        e_v = parse(existing[start_loc])
    else:
        e_v = existing[start_loc]

    if row[start_loc].value is not None and existing[start_loc] != 'None':
	if r_v < e_v:
            to_return = True

    return to_return



def get_uid(row, sheet):
    """return a row's UID based on criteria"""
    vals = ["Implementing agency", "Local partner agency" , "District", 
            "VDC / Municipalities", "Municipal Ward", "Action type", 
            "Action description", "# Items / # Man-hours / NPR",
            "Total Number Households"]
    key = ""

    for v in vals:
        try:
            key += xstr(row[column_index_from_string(find_in_header(sheet, v))-1].value)
        except:
            print 'broken!! ' + str(sheet)

    return key

def get_values(r):
    """returns values of a row or a column"""
    ret = []
    for v in r:
        ret.append(xstr(v.value))

    return ret

def send_wb(path, wb, src):
    print 'Sending: ' + path
    if src == 'db':
        client.put_file(path, wrtex.save_virtual_workbook(wb))

    elif src == 'local':
        print 'path is ' + path
        print 'splt ' + path.rsplit('/', 1)[0]
        if not os.path.exists(path.rsplit('/', 1)[0]):
            print '**Doesnt exist'
            os.makedirs(path.rsplit('/', 1)[0])
        wb.save(path)



def wb_format(wb):
    """check to see if a report is correct and in the new report format"""
    must_contain = ['Distributions', 'Training', 'Reference']

    match_count = 0
    for s in wb.worksheets:
        if s.title in must_contain:
            match_count+=1

    if match_count < 3:
        return False
    else:
        return True

def clean_file(wb, path, src):
    """cycle through a report and apply cleaning algorithms"""
    
    #get our two sheets
    db = wb.get_sheet_by_name('Distributions')
    ref = wb.get_sheet_by_name('Reference')

    #setup log
    rname =  path.rsplit('/', 1)[1]
    report_line = '******Report for ' + rname + ' ******'
    report_a_log(report_line, rname)


    #####do edit stuff
    #algos return db, ref, message

    #algo1
    db, ref, message = clean.algo1(db,ref) 
    report_a_log(message, rname)

    #algo2
    db, ref, message = clean.algo2(db,ref)
    report_a_log(message, rname)

    #algo3
#    db, ref, message = clean.algo3(db,ref)
#    report_a_log(message, rname)

    #algo4
    db, ref, message = clean.algo4(db,ref)
    report_a_log(message, rname)

    #algo5
    db, ref, message = clean.algo5(db,ref)
    report_a_log(message, rname)

    #algo6
    db, ref, message = clean.algo6(db,ref)
    report_a_log(message, rname)

    #algo7
    db, ref, message = clean.algo7(db,ref)
    report_a_log(message, rname)

    #algo8
    db, ref, message = clean.algo8(db,ref)
    report_a_log(message, rname)

    #algo9
    db, ref, message = clean.algo9(db,ref)
    report_a_log(message, rname)

    #algo10
    db, ref, message = clean.algo10(db,ref)
    report_a_log(message, rname)

    #algo11
    db, ref, message = clean.algo11(db,ref)
    report_a_log(message, rname)

    #algo12
    db, ref, message = clean.algo12(db,ref)
    report_a_log(message, rname)

    #algo13
    db, ref, message = clean.algo13(db,ref)
    report_a_log(message, rname)

    #algo14
    db, ref, message = clean.algo14(db,ref)
    report_a_log(message, rname)

    #algo15
    db, ref, message = clean.algo15(db,ref)
    report_a_log(message, rname)

    #algo16
    db, ref, message = clean.algo16(db,ref)
    report_a_log(message, rname)

    #algo17
    db, ref, message = clean.algo17(db,ref)
    report_a_log(message, rname)

    #algo18
    db, ref, message = clean.algo18(db,ref)
    report_a_log(message, rname)

    #algo19
    db, ref, message = clean.algo19(db,ref)
    report_a_log(message, rname)


    #dummy empty log to send to finalize logging
    report_a_log(' ','text')

    #upload with name of file at end
    #we need to upload the new version!!!!!!!!
    send_wb(path.rsplit('/', 1)[0] + '/edited/' + path.rsplit('/', 1)[1], wb, src)
    print 'uploaded! ' + path.rsplit('/', 1)[0] + '/edited/' + path.rsplit('/', 1)[1]

report_recvd = False
current_path = ''
current_log = []

def report_a_log(log_value, path):
    """write out contents for a given log - creates new entry if a new path is given"""
    #todo: this is gross
    global report_recvd
    global current_path
    global current_log


    #if module is starting and we haven't logged anything
    if not report_recvd:
        current_path = path
        report_recvd = True
    

    #if we are recieving a new path
    elif current_path != path:
        current_path = path
        
        #write out
        with open('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl/logs/cleaned_log'+
            time.strftime("%m-%d-%y_%H:%M_%S") +'.txt', 'w') as f:
            for log in current_log:
                f.write(str(log)+'\n')
        f.close()

        #move on to next log
        current_log.append('')
        current_log.append('')

    current_log.append(log_value)
    current_log.append('')



def find_in_header(sheet, find_val):
    """find the coordinate of a value in header (assumes header is in row 1)"""
    for row in sheet.iter_rows('A1:' + find_last_value(sheet,'A','r')):
        for cell in row:
            if cell.value == find_val:
                return cell.column

    #if we haven't returned anything yet
    print 'No header found for: ' + find_val
    return None

def colvals_notincol(sheet_val,col_val,sheet_ref,col_ref):
    """return values from a column that are NOT in a reference column"""
    not_in = []
    to_search = []

    #create an array from sheet_ref with values to be searched (as opposed to nested loops)
    #iter_rows syntax: sheet.iter_rows('A1:A2')
    for row in sheet_ref.iter_rows(col_ref + "2:" + 
        find_last_value(sheet_ref, col_ref, 'c')):

        for cell in row:               
                try:
                    to_search.append(xstr(cell.value))
                except:
                    to_search.append(str(cell.value))

    #now search through vals and see if they're present
    for row in sheet_val.iter_rows(col_val + "2:" + 
        find_last_value(sheet_val, col_val, 'c')):

        for cell in row:
            if xstr(cell.value) not in to_search:
                try:
                    not_in.append(xstr(cell.value))
                except:
                    not_in.append(str(cell.value))

    return not_in
             


def find_last_value(sheet, start_location, r_or_c):
    """find position of last value in a given row or column"""
    #extract form r_or_c if we should be iterating a row or column
    
    row_it = 0
    col_it = 0
    if r_or_c == 'c':
        row_it = 1
    elif r_or_c == 'r':
        col_it = 1
    else:
        raise Exception("r_or_c must be r or c!")
    
    #look for a cell without value, and if we find a blank, traverse 100 more
    #to make sure there's no blanks
    blank_count = 0
    found = False
    cur_cell = sheet[start_location+'1']
    while not found:
        cur_cell = cur_cell.offset(row = row_it, column = col_it)

        if not cur_cell.value:
            if blank_count == 0:
                #coord of a cell step back 1
                last_found = cur_cell.offset(row = -row_it, column= -col_it).coordinate
                blank_count+=1

            elif blank_count == 100:
                found = True
            else:
                blank_count+=1
        else:
            blank_count=0

    return last_found


def pull_wb(location, src):
    """return an excel file from either local or source"""

    return wb_strip(location, src)

def wb_strip(location, src):
    if src == 'db':
        w = load_workbook(pull_from_db(location), read_only = True, data_only = True)

    else:
        w = load_workbook(location, read_only = True, data_only = True)

    new_wb = Workbook()
    for v in w.worksheets:
        cur_w = new_wb.create_sheet(1, v.title)
        for r in v.rows:
            for v in r:
                if v.value is not None:
                    cur_w[v.coordinate] = v.value

    print "Pulled: " + location
    print "With tabs: " + str(new_wb.get_sheet_names())

    return new_wb

def pull_from_db(location):
    """pull a file from dropbox"""
    to_ret = cStringIO.StringIO()

    with client.get_file(location) as f:
        to_ret.write(f.read())
    f.close()

    return to_ret

def get_file_list(path, src):
    """return file list from local or db"""
    if src == 'db':
        meta = client.metadata(path, list=True)
        return [str(f['path']) for f in meta['contents'] if re.search('xls|xlsx$',str(f))]

    elif src == 'local':
        return [str(path +'/' + f) for f in listdir(path) if isfile(join(path,f)) and re.search('xls|xlsx$',str(f))]


def xstr(conv):
    """return a properly encoded string"""
    try:
        return conv.encode('utf8')
    except:
        return str(conv)

def test():
    return load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl/clean_test.xlsx', data_only=True)

if __name__ == '__main__':
    iterate_reports()
