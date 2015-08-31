import unittest
import etl
from openpyxl import load_workbook
from openpyxl import Workbook

import os 


sample_wb_new_format = load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl/clean_test.xlsx', data_only = True)
sample_wb = load_workbook('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl_test.xlsx')
db = sample_wb.get_sheet_by_name('Database')
ref = sample_wb.get_sheet_by_name('Reference')

class TestEtl(unittest.TestCase):

    def test_wb_format_false(self):
        self.assertEqual(etl.wb_format(sample_wb), False)

    def test_wb_format_true(self):
        self.assertEqual(etl.wb_format(sample_wb_new_format), True)

    def test_find_last_value_row(self):
        self.assertEqual(etl.find_last_value(ref, 'A', 'r'), 'Z1')

    def test_find_last_value_column(self):
        self.assertEqual(etl.find_last_value(ref, 'Z', 'c'), 'Z11')

    def test_find_in_header(self):
        self.assertEqual(etl.find_in_header(ref, 'TESTINGCOL'), 'J')       

    def test_colvals_notincol(self):
        self.assertEqual(tuple(etl.colvals_notincol(db, 'A', ref, 'A')), 
            tuple(['notincluded1','notincluded2','notincluded3']))

    def test_report_a_log(self):
        etl.report_a_log('**FILE 1**', 'file1')
        etl.report_a_log('ent2', 'file1')
        etl.report_a_log('ent2', 'file1')
        etl.report_a_log('**FILE 2**', 'file2')
        etl.report_a_log('ent5', 'file2')
        etl.report_a_log('ent5', 'file2')
        etl.report_a_log('**FILE 3**', 'file3')
        etl.report_a_log('ent5', 'file3')
        etl.report_a_log('ent5', 'file3')

        etl.report_a_log('', 'text')

        assert os.path.exists('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl/logs/cleaned_log.txt')

    def test_consolidate(self):
        #create historical db
        db = Workbook().active
        db.append(("Implementing agency", "Local partner agency" , "District", 
            "VDC / Municipalities", "Municipal Ward", "Action type", 
            "Action description", "# Items / # Man-hours / NPR",
            "Total Number Households"))
        db.append(('val', 'key', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        db.append(('dup1', 'dupkey1', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        db.append(('row2val', 'key1', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        db.append(('row3val', 'key2', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        db.append(('dup2', 'dupkey2', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))

        #create two sheets to add
        wb1 = Workbook()
        wb1.create_sheet(2, 'Distributions')
        ws1 = wb1.get_sheet_by_name('Distributions')
        ws1.append(("Implementing agency", "Local partner agency" , "District",
            "VDC / Municipalities", "Municipal Ward", "Action type", 
            "Action description", "# Items / # Man-hours / NPR",
            "Total Number Households"))
        ws1.append(('val', 'key', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        ws1.append(('dup1', 'dupkey1', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        ws1.append(('notdupedws1', 'notdupedvalws1', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))

        wb2 = Workbook()
        wb2.create_sheet(1, 'Distributions')
        ws2 = wb2.get_sheet_by_name('Distributions')
        ws2.title = 'Distributions'
        ws2.append(("Implementing agency", "Local partner agency" , "District", 
            "VDC / Municipalities", "Municipal Ward", "Action type", 
            "Action description", "# Items / # Man-hours / NPR",
            "Total Number Households"))
        ws2.append(('val', 'key', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        ws2.append(('dup2', 'dupkey2', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))
        ws2.append(('notdupedws2', 'notdupedvalws2', 'None', 'None', 'None', 'None', 'None', 'None', 'None', 'Last'))

        #weird bug
        wb2.get_sheet_by_name('Distributions1').title = 'Distributions'
        print 'in test: ' + str(wb1.get_sheet_names())
        print 'in test2: ' + str(wb2.get_sheet_names())

        cons = etl.consolidate(db, (wb1, wb2), 'J')
        cons_sheet = cons.get_sheet_by_name('Consolidated')

        self.assertEqual(set(etl.get_values(cons_sheet.columns[9])), 
            (set(['None', 'valkeyNoneNoneNoneNoneNoneNoneNone','row2valkey1NoneNoneNoneNoneNoneNoneNone',
                'row3valkey2NoneNoneNoneNoneNoneNoneNone','dup1dupkey1NoneNoneNoneNoneNoneNoneNone',
                'notdupedws1notdupedvalws1NoneNoneNoneNoneNoneNoneNone','dup2dupkey2NoneNoneNoneNoneNoneNoneNone',
                'notdupedws2notdupedvalws2NoneNoneNoneNoneNoneNoneNone'])))

    def test_get_values(self):
        db = Workbook().active
        db.append(('val', 'key'))
        self.assertEqual(etl.get_values(db.rows[0]),['val','key'])

    def test_keep_dict(self):
        """row is new value, dict is old"""
        db = Workbook().active
        db.append(('Activity Status', 'Completion Date\n (Actual or Planned)', 'Start date \n(Actual or Planned)'))
        db.append(('Completed', '12/15/2015','12/10/2015'))
        d_v =  ('Completed', '12/15/2015','12/12/2015')
        self.assertTrue(etl.keep_dict(db.rows[1], d_v, db))

    def test_none_row_true(self):
        self.assertTrue(etl.none_row('NoneNoneNoneNoneNoneNone'))

    def test_none_row_false(self):
        self.assertFalse(etl.none_row('thisisnotrepeating'))

if __name__ == '__main__':
    unittest.main()
