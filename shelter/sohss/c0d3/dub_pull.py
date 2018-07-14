import os
import json
import traceback
from openpyxl import (
    Workbook,
    load_workbook as _load_workbook,
)
from openpyxl.utils import column_index_from_string


"""
- Current File structure
.
├── 4W overview.xlsx                            (mapping file)
├── main.py
├── outputwithfilename.xlsx                     (output with filename)
├── output.xlsx                                 (output)
└── Ws                                          (input files)
    ├── 120524_3w_all_clusters_april_public.xlsx
    ├─....


- requirement:
openpyxl==2.5.4

"""

FILES_PREFIX = './Ws'
XLS_EXTENSIONS = ['', '.xlsx', '.xlsm']
map_filename = './4W overview.xlsx'
map_sheets = ['Kriti', 'Nagma', 'Shweta']

MAPPING_COLUMN_META = [
    ['Date Reported', 'C'],
    ['Donor', 'D'],
    ['Organisation', 'E'],
    ['Implementing Partner', 'F'],
    ['Activity category', 'G'],
    ['Area of activities', 'H'],
    ['Activity detail', 'I'],
    ['Admin level 1', 'J'],
    ['Admin level 1 code', 'K'],
    ['Admin level 2', 'L'],
    ['Admin level 2 code', 'M'],
    ['Status', 'N'],
    ['# of HH reached', 'O'],
    ['Number of beneficiaries reached', 'P'],
    ['Start date', 'Q'],
    ['End date', 'R'],
]


def clean_filename(filename):
    _filename = filename.lstrip()
    _filename = _filename.replace('\n', '')
    return _filename


def gen_file_path(filename):
    for extension in XLS_EXTENSIONS:
        _filename = os.path.join(
            FILES_PREFIX,
            clean_filename(filename + extension)
        )
        if os.path.isfile(_filename):
            return _filename
    return filename


def load_workbook(filename, *args, **kwargs):
    return _load_workbook(gen_file_path(filename), *args, **kwargs)


def get_column_index(column_in_string):
    return column_index_from_string(str(column_in_string.replace(' ', ''))) - 1


def get_cell_value(row, column):
    columns = column.split(',')
    if len(columns) == 1:
        try:
            return row[get_column_index(column)].value
        except IndexError:
            return None
    values = []
    for column in columns:
        try:
            values.append(int(get_cell_value(row, column)))
        except (TypeError, ValueError):
            pass
    return sum(values)


def print_warning(*args, **kwargs):
    print(*args, **kwargs)


def print_error(*args, **kwargs):
    print('*' * 50)
    print(*args, **kwargs)
    print('-' * 50)


def print_meta(row):
    print(json.dumps([
        {'type': map[0], 'column': row[get_column_index(map[1])].value}
        for map in MAPPING_COLUMN_META
    ]))


def is_row_empty(row, columns):
    for column in columns:
        if column is not None:
            value = get_cell_value(row, column)
            if value is not None:
                return False
    return True


class ShelterClusterExtract():
    def __init__(self, add_filename=False):
        # Mapping File
        self.mwb = _load_workbook(
            filename=map_filename, data_only=True, read_only=True,
        )

        # Output File
        self.add_filename = add_filename
        self.owb = Workbook()
        self.ows = self.owb.active
        self.orow = 1  # initial staring row
        self.success_files = []
        self.error_files = []
        self.files_with_hidden_cols_rows = []

        self.init_headers()

    def init_headers(self):
        column_index = 1
        if self.add_filename:
            self.ows.cell(row=self.orow, column=column_index, value='File')
            column_index += 1
        for map in MAPPING_COLUMN_META:
            self.ows.cell(row=self.orow, column=column_index, value=map[0])
            column_index += 1
        self.orow += 1

    def gather_hidden_cols_rows(self, sheet, path, sheet_name, user):
        hidden_rows = []
        hidden_columns = []
        for row in sheet.row_dimensions:
            if sheet.row_dimensions[row].hidden:
                hidden_rows.append(row)
        for col in sheet.column_dimensions:
            if sheet.column_dimensions[col].hidden:
                hidden_columns.append(col)
        if len(hidden_rows) > 0 or len(hidden_columns) > 0:
            print('\nFound: ', path, '\n')
            self.files_with_hidden_cols_rows.append({
                'user': user,
                'sheet': sheet_name,
                'path': path,
                'rows': hidden_rows,
                'columns': hidden_columns,
            })
        else:
            print('.', end='')

    def show_summary(self):
        print('DONE')
        print('*' * 22)
        print('Files with success: ', len(self.success_files))
        [
            print(file['sheet'], ':', file['path'])
            for file in self.success_files
        ]
        print('-' * 22)
        print('Files with error: ', len(self.error_files))
        [
            print(
                file['sheet'], ':', file['path'], ' --> ', file['error'], '\n',
            )
            for file in self.error_files
        ]
        print(
            'Files with hidden rows/columns: ',
            len(self.files_with_hidden_cols_rows)
        )
        """
        [
            print(
                '*' * 22, file['user'], ':', file['path'], ':',
                file['sheet'], '*' * 22, '\n rows: ', file['rows'],
                '\n cols: ', file['columns'], '\n'
            )
            for file in self.files_with_hidden_cols_rows
        ]
        """

    def write_to_output(self, row, columns, filename):
        ocolumn = 1
        if self.add_filename:
            self.ows.cell(
                row=self.orow, column=ocolumn,
                value=filename,
            )
            ocolumn = 2
        if is_row_empty(row, columns):
            return
        for column in columns:
            if column is not None:
                self.ows.cell(
                    row=self.orow, column=ocolumn,
                    value=get_cell_value(row, column),
                )
            ocolumn += 1
        self.orow += 1

    def extract_document(self, mrow, user):
        # Input File
        filename = get_cell_value(mrow, 'A')
        try:
            sheet_name = get_cell_value(mrow, 'S')
        except Exception:
            print_warning('Warning: no sheet name specified')
            sheet_name = None

        header = get_cell_value(mrow, 'B')
        columns = [
            get_cell_value(mrow, map[1])
            for map in MAPPING_COLUMN_META
        ]

        # print_meta(mrow)

        rwb = load_workbook(
            filename=filename, data_only=True,  # read_only=True
        )
        rws = rwb[sheet_name] if sheet_name else rwb.active
        # print(rwb.get_sheet_names())

        # self.gather_hidden_cols_rows(rws, filename, sheet_name, user)
        for row in rws.iter_rows(row_offset=int(header)):
            self.write_to_output(row, columns, filename)

    def run(self, sheet):
        for mrow in self.mwb[sheet].iter_rows(row_offset=1):
            if is_row_empty(mrow, ['A']):
                continue
            try:
                self.extract_document(mrow, sheet)
                self.success_files.append({
                    'path': gen_file_path(mrow[0].value), 'sheet': sheet,
                })
            except Exception as e:
                print('Error Occured for file: ', gen_file_path(mrow[0].value))
                self.error_files.append({
                    'path': gen_file_path(mrow[0].value), 'error': e,
                    'sheet': sheet,
                })
                print_error(traceback.format_exc())

    def start(self, output_filename):
        self.mwss = []
        for map_sheet in map_sheets:
            self.run(map_sheet)
        self.owb.save(output_filename)
        self.show_summary()


if __name__ == '__main__':
    ShelterClusterExtract(add_filename=True).start('outputwithfilename.xlsx')
    ShelterClusterExtract(add_filename=False).start('output.xlsx')import os
import json
import traceback
from openpyxl import (
    Workbook,
    load_workbook as _load_workbook,
)
from openpyxl.utils import column_index_from_string


"""
- Current File structure
.
├── 4W overview.xlsx                            (mapping file)
├── main.py
├── outputwithfilename.xlsx                     (output with filename)
├── output.xlsx                                 (output)
└── Ws                                          (input files)
    ├── 120524_3w_all_clusters_april_public.xlsx
    ├─....


- requirement:
openpyxl==2.5.4

"""

FILES_PREFIX = './Ws'
XLS_EXTENSIONS = ['', '.xlsx', '.xlsm']
map_filename = './4W overview.xlsx'
map_sheets = ['Kriti', 'Nagma', 'Shweta']

MAPPING_COLUMN_META = [
    ['Date Reported', 'C'],
    ['Donor', 'D'],
    ['Organisation', 'E'],
    ['Implementing Partner', 'F'],
    ['Activity category', 'G'],
    ['Area of activities', 'H'],
    ['Activity detail', 'I'],
    ['Admin level 1', 'J'],
    ['Admin level 1 code', 'K'],
    ['Admin level 2', 'L'],
    ['Admin level 2 code', 'M'],
    ['Status', 'N'],
    ['# of HH reached', 'O'],
    ['Number of beneficiaries reached', 'P'],
    ['Start date', 'Q'],
    ['End date', 'R'],
]


def clean_filename(filename):
    _filename = filename.lstrip()
    _filename = _filename.replace('\n', '')
    return _filename


def gen_file_path(filename):
    for extension in XLS_EXTENSIONS:
        _filename = os.path.join(
            FILES_PREFIX,
            clean_filename(filename + extension)
        )
        if os.path.isfile(_filename):
            return _filename
    return filename


def load_workbook(filename, *args, **kwargs):
    return _load_workbook(gen_file_path(filename), *args, **kwargs)


def get_column_index(column_in_string):
    return column_index_from_string(str(column_in_string.replace(' ', ''))) - 1


def get_cell_value(row, column):
    columns = column.split(',')
    if len(columns) == 1:
        try:
            return row[get_column_index(column)].value
        except IndexError:
            return None
    values = []
    for column in columns:
        try:
            values.append(int(get_cell_value(row, column)))
        except (TypeError, ValueError):
            pass
    return sum(values)


def print_warning(*args, **kwargs):
    print(*args, **kwargs)


def print_error(*args, **kwargs):
    print('*' * 50)
    print(*args, **kwargs)
    print('-' * 50)


def print_meta(row):
    print(json.dumps([
        {'type': map[0], 'column': row[get_column_index(map[1])].value}
        for map in MAPPING_COLUMN_META
    ]))


def is_row_empty(row, columns):
    for column in columns:
        if column is not None:
            value = get_cell_value(row, column)
            if value is not None:
                return False
    return True


class ShelterClusterExtract():
    def __init__(self, add_filename=False):
        # Mapping File
        self.mwb = _load_workbook(
            filename=map_filename, data_only=True, read_only=True,
        )

        # Output File
        self.add_filename = add_filename
        self.owb = Workbook()
        self.ows = self.owb.active
        self.orow = 1  # initial staring row
        self.success_files = []
        self.error_files = []
        self.files_with_hidden_cols_rows = []

        self.init_headers()

    def init_headers(self):
        column_index = 1
        if self.add_filename:
            self.ows.cell(row=self.orow, column=column_index, value='File')
            column_index += 1
        for map in MAPPING_COLUMN_META:
            self.ows.cell(row=self.orow, column=column_index, value=map[0])
            column_index += 1
        self.orow += 1

    def gather_hidden_cols_rows(self, sheet, path, sheet_name, user):
        hidden_rows = []
        hidden_columns = []
        for row in sheet.row_dimensions:
            if sheet.row_dimensions[row].hidden:
                hidden_rows.append(row)
        for col in sheet.column_dimensions:
            if sheet.column_dimensions[col].hidden:
                hidden_columns.append(col)
        if len(hidden_rows) > 0 or len(hidden_columns) > 0:
            print('\nFound: ', path, '\n')
            self.files_with_hidden_cols_rows.append({
                'user': user,
                'sheet': sheet_name,
                'path': path,
                'rows': hidden_rows,
                'columns': hidden_columns,
            })
        else:
            print('.', end='')

    def show_summary(self):
        print('DONE')
        print('*' * 22)
        print('Files with success: ', len(self.success_files))
        [
            print(file['sheet'], ':', file['path'])
            for file in self.success_files
        ]
        print('-' * 22)
        print('Files with error: ', len(self.error_files))
        [
            print(
                file['sheet'], ':', file['path'], ' --> ', file['error'], '\n',
            )
            for file in self.error_files
        ]
        print(
            'Files with hidden rows/columns: ',
            len(self.files_with_hidden_cols_rows)
        )
        """
        [
            print(
                '*' * 22, file['user'], ':', file['path'], ':',
                file['sheet'], '*' * 22, '\n rows: ', file['rows'],
                '\n cols: ', file['columns'], '\n'
            )
            for file in self.files_with_hidden_cols_rows
        ]
        """

    def write_to_output(self, row, columns, filename):
        ocolumn = 1
        if self.add_filename:
            self.ows.cell(
                row=self.orow, column=ocolumn,
                value=filename,
            )
            ocolumn = 2
        if is_row_empty(row, columns):
            return
        for column in columns:
            if column is not None:
                self.ows.cell(
                    row=self.orow, column=ocolumn,
                    value=get_cell_value(row, column),
                )
            ocolumn += 1
        self.orow += 1

    def extract_document(self, mrow, user):
        # Input File
        filename = get_cell_value(mrow, 'A')
        try:
            sheet_name = get_cell_value(mrow, 'S')
        except Exception:
            print_warning('Warning: no sheet name specified')
            sheet_name = None

        header = get_cell_value(mrow, 'B')
        columns = [
            get_cell_value(mrow, map[1])
            for map in MAPPING_COLUMN_META
        ]

        # print_meta(mrow)

        rwb = load_workbook(
            filename=filename, data_only=True,  # read_only=True
        )
        rws = rwb[sheet_name] if sheet_name else rwb.active
        # print(rwb.get_sheet_names())

        # self.gather_hidden_cols_rows(rws, filename, sheet_name, user)
        for row in rws.iter_rows(row_offset=int(header)):
            self.write_to_output(row, columns, filename)

    def run(self, sheet):
        for mrow in self.mwb[sheet].iter_rows(row_offset=1):
            if is_row_empty(mrow, ['A']):
                continue
            try:
                self.extract_document(mrow, sheet)
                self.success_files.append({
                    'path': gen_file_path(mrow[0].value), 'sheet': sheet,
                })
            except Exception as e:
                print('Error Occured for file: ', gen_file_path(mrow[0].value))
                self.error_files.append({
                    'path': gen_file_path(mrow[0].value), 'error': e,
                    'sheet': sheet,
                })
                print_error(traceback.format_exc())

    def start(self, output_filename):
        self.mwss = []
        for map_sheet in map_sheets:
            self.run(map_sheet)
        self.owb.save(output_filename)
        self.show_summary()


if __name__ == '__main__':
    ShelterClusterExtract(add_filename=True).start('../d0cz/outputwithfilename.xlsx')
    ShelterClusterExtract(add_filename=False).start('../d0cz/output.xlsx')

