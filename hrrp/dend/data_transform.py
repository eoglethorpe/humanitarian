import csv
import re
from os import listdir
from os.path import isfile, join

from openpyxl import load_workbook

CENT = 'start'
PATH = './data/xls/'
FILE = '4W-National-R8-160502.xlsx'


def head(inp, data):
    """add in the district points, needed for when we add POs"""
    inp.append(['id', 'value'])
    inp.append([CENT, '1'])
    for v in list(set(v[0] for v in data)):
        inp.append([CENT + '.' + v, 1])

    return inp


def check_cols(inp):
    #check to see if right columns. possible the file doesn't have a SN column, so add if not
    if inp[1][1].value != 'District' or inp[1][4].value != 'Partner Organisation':
        if inp[1][0].value == 'District' and inp[1][3].value == 'Partner Organisation':
            return (0,3)
        else:
            raise ValueError('Cols not equal')

    else:
        return (1,4)


def add_dist_po(inp, data):
    for v in data:
        inp.append(['{0}.{1}.{2}'.format(CENT, v[0], v[1]),1])

    return inp


def gen_tupz(sht):
    tupz = []
    pass_first = True

    col_adds = check_cols(sht)
    for r in sht.iter_rows():
        if pass_first:
            pass_first = False

        else:
            cur_tup = (r[col_adds[0]].value, r[col_adds[1]].value)
            if cur_tup not in tupz:
                tupz.append(cur_tup)

    return tupz


def export(out, cur_sht):
    with open('./data/xls/out/{0}_out.csv'.format(cur_sht.split('.')[0]), 'w') as f:
        w = csv.writer(f)
        for r in out:
            w.writerow(r)


def main():
    exported = []
    p = re.compile(r'(\d{6})')
    for cur_sht in [f for f in listdir(PATH) if isfile(join(PATH, f)) and f[0] not in ['.', '~']]:

        match = re.findall(p, cur_sht)
        if len(match) != 1:
            Exception('no date!')

        #check to see if yearmonth (1710 ie) is already inserted. we only want one entry per month
        elif match[0][0:4] in exported:
            print('passing ' + cur_sht)
            pass

        else:
            print('doing ' + cur_sht)
            exported.append(match[0][0:4])
            sht = load_workbook(PATH + cur_sht).get_sheet_by_name('4W')
            #make sure dist and PO cols are same lengths
            assert (len(sht['A']) == len(sht['D']))

            data = gen_tupz(sht)

            out = []
            out_s = head(out, data)
            out_s = add_dist_po(out, data)
            export(out, cur_sht)

if __name__ == '__main__':
    main()
